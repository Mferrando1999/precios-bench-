#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
actualiza_precios.py
Lee "bench automatizacion.xlsx", visita cada URL de la columna Links (F),
extrae precio+moneda y escribe el precio convertido (según tipo de cambio en filas 3-4)
en la columna de la semana indicada (ej. week 8 = columna AY).

Uso (local):
  pip install openpyxl requests beautifulsoup4 lxml
  python actualiza_precios.py --input "bench automatizacion.xlsx" --output "bench automatizacion_actualizado.xlsx" --week 8

Opcional (si algunas webs requieren JS):
  pip install playwright
  playwright install chromium
  python actualiza_precios.py --input ... --output ... --week 8 --use-playwright
"""
import argparse
import datetime as dt
import json
import re
from decimal import Decimal, InvalidOperation
from typing import Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

import requests
from bs4 import BeautifulSoup


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/123.0.0.0 Safari/537.36"
)


CURRENCY_SYMBOLS = {
    "€": "EUR",
    "£": "GBP",
    "$": "USD",
    "¥": "JPY",
    "₩": "KRW",
    "C$": "CAD",
    "A$": "AUD",
}


def _to_decimal(x) -> Optional[Decimal]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        try:
            return Decimal(str(x))
        except InvalidOperation:
            return None
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return None
        # Normaliza separadores (1.234,56) vs (1,234.56)
        s = s.replace("\xa0", " ")
        s = re.sub(r"[^\d,.\-]", "", s)
        if not s:
            return None

        # heurística: si hay coma y punto, asumimos que el último es decimal
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        else:
            # si solo coma, la tratamos como decimal
            if "," in s and "." not in s:
                s = s.replace(",", ".")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    return None


def _extract_currency_from_text(text: str) -> Optional[str]:
    if not text:
        return None
    for sym, cur in CURRENCY_SYMBOLS.items():
        if sym in text:
            return cur
    m = re.search(r"\b(EUR|USD|GBP|CAD|AUD|JPY|KRW)\b", text, re.I)
    if m:
        return m.group(1).upper()
    return None


def _pick_first_jsonld(soup: BeautifulSoup) -> list:
    """Devuelve una lista de objetos JSON-LD parseados (puede ser vacía)."""
    out = []
    for tag in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = tag.string or tag.get_text()
        if not raw:
            continue
        raw = raw.strip()
        try:
            data = json.loads(raw)
            out.append(data)
        except Exception:
            # algunos sitios meten múltiple JSON en el mismo script; intentamos una extracción simple
            try:
                raw2 = re.sub(r"[\n\r\t]", " ", raw)
                # intenta capturar el primer objeto JSON
                m = re.search(r"(\{.*\})", raw2)
                if m:
                    out.append(json.loads(m.group(1)))
            except Exception:
                continue
    return out


def _find_offer_in_jsonld(obj) -> Optional[dict]:
    """Busca offer con price/priceCurrency dentro de JSON-LD."""
    if obj is None:
        return None

    if isinstance(obj, list):
        for it in obj:
            got = _find_offer_in_jsonld(it)
            if got:
                return got
        return None

    if isinstance(obj, dict):
        # Product -> offers
        if obj.get("@type") in ("Product", ["Product"]):
            offers = obj.get("offers")
            if offers:
                if isinstance(offers, list):
                    for o in offers:
                        if isinstance(o, dict) and (o.get("price") or o.get("lowPrice")):
                            return o
                if isinstance(offers, dict) and (offers.get("price") or offers.get("lowPrice")):
                    return offers

        # graph
        if "@graph" in obj:
            return _find_offer_in_jsonld(obj["@graph"])

        # recorrer claves
        for v in obj.values():
            got = _find_offer_in_jsonld(v)
            if got:
                return got

    return None


def extract_price_currency_html(html: str) -> Tuple[Optional[Decimal], Optional[str], str]:
    """
    Intenta extraer price+currency de HTML sin JS.
    Devuelve (precio, moneda, fuente).
    """
    soup = BeautifulSoup(html, "lxml")

    # 1) JSON-LD
    jsonlds = _pick_first_jsonld(soup)
    offer = _find_offer_in_jsonld(jsonlds)
    if offer:
        price = _to_decimal(offer.get("price") or offer.get("lowPrice"))
        currency = (offer.get("priceCurrency") or offer.get("pricecurrency"))
        if isinstance(currency, str):
            currency = currency.upper().strip()
        if price is not None and currency:
            return price, currency, "jsonld"

    # 2) Meta tags comunes (OG / product)
    meta_amount = soup.find("meta", attrs={"property": "product:price:amount"}) or soup.find(
        "meta", attrs={"name": "product:price:amount"}
    )
    meta_currency = soup.find("meta", attrs={"property": "product:price:currency"}) or soup.find(
        "meta", attrs={"name": "product:price:currency"}
    )
    if meta_amount and meta_amount.get("content"):
        price = _to_decimal(meta_amount.get("content"))
        currency = meta_currency.get("content").upper().strip() if meta_currency and meta_currency.get("content") else None
        if price is not None and currency:
            return price, currency, "meta_product_price"

    # 3) Shopify: meta itemprop price / priceCurrency
    item_price = soup.find(attrs={"itemprop": "price"})
    item_currency = soup.find(attrs={"itemprop": "priceCurrency"})
    if item_price and item_price.get("content"):
        price = _to_decimal(item_price.get("content"))
        currency = item_currency.get("content").upper().strip() if item_currency and item_currency.get("content") else None
        if price is not None and currency:
            return price, currency, "microdata_itemprop"

    # 4) Heurística: buscar en texto algo que parezca precio
    text = soup.get_text(" ", strip=True)
    # patrones: £1,234.56 / 1,234.56 USD / EUR 123.45
    patterns = [
        r"([€£$]\s?\d[\d\s.,]{0,12}\d)",
        r"\b(EUR|USD|GBP)\s?\d[\d\s.,]{0,12}\d\b",
        r"\b\d[\d\s.,]{0,12}\d\s?(EUR|USD|GBP)\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            s = m.group(0)
            currency = _extract_currency_from_text(s) or _extract_currency_from_text(text[:300])
            price = _to_decimal(s)
            if price is not None and currency:
                return price, currency, "text_heuristic"

    return None, None, "not_found"


def fetch_html_requests(url: str, timeout: int = 25) -> str:
    headers = {"User-Agent": USER_AGENT, "Accept-Language": "en-US,en;q=0.9,es;q=0.8"}
    r = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    return r.text


async def fetch_html_playwright(url: str, timeout_ms: int = 35000) -> str:
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9,es;q=0.8"})
        await page.goto(url, timeout=timeout_ms, wait_until="networkidle")
        html = await page.content()
        await browser.close()
        return html


def find_week_column(ws, week: int) -> int:
    """Busca el número de semana en la fila 6 y devuelve el índice de columna."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=6, column=c).value == week:
            return c
    raise ValueError(f"No encontré la week {week} en la fila 6.")


def read_fx_for_column(ws, col: int) -> Tuple[Optional[str], Optional[Decimal]]:
    """
    En fila 3: '1USD=' / '1GBP=' etc.
    En fila 4: tasa numérica.
    Devuelve (moneda_base, tasa). Interpretación: 1 <moneda_base> = <tasa> EUR (o tu moneda destino).
    """
    label = ws.cell(row=3, column=col).value
    rate = ws.cell(row=4, column=col).value
    base = None
    if isinstance(label, str):
        m = re.search(r"1\s*([A-Z]{3})\s*=", label.strip().upper())
        if m:
            base = m.group(1)
    return base, _to_decimal(rate)


def ensure_log_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    name = "LOG"
    if name in wb.sheetnames:
        return wb[name]
    wslog = wb.create_sheet(name)
    wslog.append(["timestamp", "row", "url", "price_raw", "currency_raw", "fx_base", "fx_rate", "written_col", "written_value", "source", "status", "error"])
    return wslog


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Ruta al xlsx de entrada")
    ap.add_argument("--output", required=True, help="Ruta al xlsx de salida")
    ap.add_argument("--week", required=True, type=int, help="Número de week a actualizar (ej. 8)")
    ap.add_argument("--use-playwright", action="store_true", help="Usa navegador headless para sitios con JS")
    ap.add_argument("--max-rows", type=int, default=0, help="0 = todas; si no, limita filas (útil para pruebas)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)
    ws = wb.active

    col_week = find_week_column(ws, args.week)
    col_letter = get_column_letter(col_week)

    fx_base, fx_rate = read_fx_for_column(ws, col_week)

    wslog = ensure_log_sheet(wb)

    start_row = 7
    link_col = 6  # F
    now = dt.datetime.now().isoformat(timespec="seconds")

    # Playwright (si aplica)
    use_playwright = bool(args.use_playwright)
    if use_playwright:
        import asyncio

    updated = 0
    errors = 0

    # Recorre filas con URL
    max_row = ws.max_row
    if args.max_rows and args.max_rows > 0:
        max_row = min(max_row, start_row + args.max_rows - 1)

    for r in range(start_row, max_row + 1):
        url = ws.cell(row=r, column=link_col).value
        if not url or not isinstance(url, str) or not url.startswith("http"):
            continue

        try:
            if use_playwright:
                html = asyncio.run(fetch_html_playwright(url))
            else:
                html = fetch_html_requests(url)

            price, currency, source = extract_price_currency_html(html)

            if price is None or not currency:
                wslog.append([now, r, url, None, None, fx_base, float(fx_rate) if fx_rate else None, col_letter, None, source, "NO_PRICE", "no se detectó precio"])
                errors += 1
                continue

            # Si hay FX configurado, y la moneda detectada coincide con la base, convertimos
            value_to_write = None
            if fx_base and fx_rate:
                if currency.upper() == fx_base.upper():
                    value_to_write = (price * fx_rate)
                else:
                    # Si el precio ya viene en EUR, lo dejamos tal cual
                    if currency.upper() == "EUR":
                        value_to_write = price
                    else:
                        wslog.append([now, r, url, float(price), currency, fx_base, float(fx_rate), col_letter, None, source, "CURRENCY_MISMATCH", f"detectado {currency} pero la columna espera {fx_base}"])
                        errors += 1
                        continue
            else:
                # Sin FX, escribe el precio detectado (sin convertir)
                value_to_write = price

            # Redondeo “humano”
            try:
                value_float = float(value_to_write.quantize(Decimal("0.01")))
            except Exception:
                value_float = float(value_to_write)

            ws.cell(row=r, column=col_week).value = value_float
            wslog.append([now, r, url, float(price), currency, fx_base, float(fx_rate) if fx_rate else None, col_letter, value_float, source, "OK", None])
            updated += 1

        except Exception as e:
            wslog.append([now, r, url, None, None, fx_base, float(fx_rate) if fx_rate else None, col_letter, None, None, "ERROR", repr(e)])
            errors += 1

    wb.save(args.output)
    print(f"Listo. Week {args.week} ({col_letter}) actualizado. Filas OK: {updated}. Con issues: {errors}. Output: {args.output}")


if __name__ == "__main__":
    main()
